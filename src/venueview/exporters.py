from __future__ import annotations

import csv
import re
from dataclasses import dataclass
from datetime import date, datetime, timedelta
from io import BytesIO, StringIO
from zipfile import ZIP_DEFLATED, ZipFile

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.page import PageMargins
from openpyxl.worksheet.table import Table, TableStyleInfo

from .models import OperationalEvent
from .pipeline import PipelineResult


EXPORT_HEADERS = (
    "Date",
    "Start",
    "End",
    "Venue",
    "Space",
    "Group",
    "Function",
    "Title",
    "Needs Review",
    "Source Count",
)

ILLEGAL_XML_CHARACTERS = re.compile(r"[\x00-\x08\x0B\x0C\x0E-\x1F]")
FORMULA_PREFIXES = ("=", "+", "-", "@")
FUNCTION_SHEET_HEADERS = (
    "Time",
    "Location",
    "Group",
    "Function",
    "Set-up",
    "HT",
    "P/M",
    "Pro",
    "Lav",
    "HH Mic",
    "Flip",
)


@dataclass(frozen=True)
class DownloadArtifact:
    data: bytes
    filename: str
    mimetype: str


def _filename_slug(value: str) -> str:
    slug = re.sub(r"[^a-z0-9]+", "-", value.casefold()).strip("-")
    return slug or "venue"


def _safe_spreadsheet_text(value: str) -> str:
    """Keep untrusted calendar text as text in CSV and Excel applications."""

    cleaned = ILLEGAL_XML_CHARACTERS.sub("", value)
    if cleaned.lstrip().startswith(FORMULA_PREFIXES):
        return f"'{cleaned}"
    return cleaned


def _download_stem(*, profile_name: str, window_start: str, window_end: str) -> str:
    return f"venueview-{_filename_slug(profile_name)}-{window_start}-to-{window_end}"


def _csv_row(event: OperationalEvent) -> list[str | int]:
    start_value = "All Day" if event.all_day else event.start.strftime("%H:%M")
    end_value = "" if event.all_day else event.end.strftime("%H:%M")
    return [
        event.local_date.isoformat(),
        start_value,
        end_value,
        _safe_spreadsheet_text(event.venue),
        _safe_spreadsheet_text(event.space),
        _safe_spreadsheet_text(event.group),
        _safe_spreadsheet_text(event.function),
        _safe_spreadsheet_text(event.title),
        _safe_spreadsheet_text("; ".join(event.needs_review)),
        event.source_count,
    ]


def _excel_row(event: OperationalEvent) -> list[object]:
    start_value: object = (
        "All Day" if event.all_day else event.start.time().replace(tzinfo=None)
    )
    end_value: object = (
        None if event.all_day else event.end.time().replace(tzinfo=None)
    )
    return [
        event.local_date,
        start_value,
        end_value,
        _safe_spreadsheet_text(event.venue),
        _safe_spreadsheet_text(event.space),
        _safe_spreadsheet_text(event.group),
        _safe_spreadsheet_text(event.function),
        _safe_spreadsheet_text(event.title),
        _safe_spreadsheet_text("; ".join(event.needs_review)),
        event.source_count,
    ]


def _clock_label(value: datetime) -> str:
    hour = value.hour % 12 or 12
    meridiem = "am" if value.hour < 12 else "pm"
    return f"{hour}:{value.minute:02d} {meridiem}"


def _time_label(event: OperationalEvent) -> str:
    if event.all_day:
        return "All Day"
    return f"{_clock_label(event.start)} - {_clock_label(event.end)}"


def _date_label(value: date, *, include_weekday: bool = False) -> str:
    prefix = f"{value.strftime('%A')}, " if include_weekday else ""
    return f"{prefix}{value.strftime('%B')} {value.day}, {value.year}"


def _csv_bytes(events: list[OperationalEvent]) -> bytes:
    stream = StringIO(newline="")
    writer = csv.writer(stream, lineterminator="\n")
    writer.writerow(EXPORT_HEADERS)
    writer.writerows(_csv_row(event) for event in events)
    return stream.getvalue().encode("utf-8-sig")


def build_csv_download(
    *,
    pipeline_result: PipelineResult,
    profile_name: str,
    mode: str,
    window_start: str,
    window_end: str,
) -> DownloadArtifact:
    stem = _download_stem(
        profile_name=profile_name,
        window_start=window_start,
        window_end=window_end,
    )
    review_rows = [event for event in pipeline_result.detailed if event.needs_review]
    if mode == "detailed":
        return DownloadArtifact(
            data=_csv_bytes(pipeline_result.detailed),
            filename=f"{stem}-detailed.csv",
            mimetype="text/csv; charset=utf-8",
        )
    if mode == "combined":
        return DownloadArtifact(
            data=_csv_bytes(pipeline_result.combined),
            filename=f"{stem}-combined.csv",
            mimetype="text/csv; charset=utf-8",
        )
    if mode != "both":
        raise ValueError("Unsupported export mode")

    stream = BytesIO()
    with ZipFile(stream, "w", compression=ZIP_DEFLATED) as archive:
        archive.writestr("detailed.csv", _csv_bytes(pipeline_result.detailed))
        archive.writestr("combined.csv", _csv_bytes(pipeline_result.combined))
        archive.writestr("review.csv", _csv_bytes(review_rows))
        archive.writestr(
            "README.txt",
            "VenueView local export\n"
            "Detailed rows preserve calendar occurrences.\n"
            "Combined rows apply configured operational rules.\n"
            "Review rows require human verification before distribution.\n",
        )
    return DownloadArtifact(
        data=stream.getvalue(),
        filename=f"{stem}-csv.zip",
        mimetype="application/zip",
    )


def _style_data_sheet(
    *, sheet, events: list[OperationalEvent], table_name: str
) -> None:
    sheet.sheet_view.showGridLines = False
    sheet.freeze_panes = "A2"
    sheet.append(list(EXPORT_HEADERS))
    for event in events:
        sheet.append(_excel_row(event))

    header_fill = PatternFill("solid", fgColor="163A5F")
    header_font = Font(color="FFFFFF", bold=True)
    light_rule = Side(style="thin", color="D9E2EC")
    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(vertical="center")
        cell.border = Border(bottom=light_rule)
    sheet.row_dimensions[1].height = 24

    for cell in sheet["A"][1:]:
        cell.number_format = "yyyy-mm-dd"
    for column in ("B", "C"):
        for cell in sheet[column][1:]:
            cell.number_format = "hh:mm"

    widths = {
        "A": 13,
        "B": 10,
        "C": 10,
        "D": 22,
        "E": 22,
        "F": 24,
        "G": 20,
        "H": 38,
        "I": 38,
        "J": 14,
    }
    for column, width in widths.items():
        sheet.column_dimensions[column].width = width
    for row in sheet.iter_rows(min_row=2, max_col=len(EXPORT_HEADERS)):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    if events:
        table = Table(
            displayName=table_name,
            ref=f"A1:J{len(events) + 1}",
        )
        table.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium2",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        sheet.add_table(table)
    else:
        sheet.auto_filter.ref = "A1:J1"
        sheet["A2"] = "No rows in this section."
        sheet["A2"].font = Font(color="5B6875", italic=True)


def _add_function_sheet(
    workbook: Workbook,
    *,
    pipeline_result: PipelineResult,
    mode: str,
    window_start: str,
    window_end: str,
) -> None:
    sheet = workbook.active
    sheet.title = "Function Sheet"
    sheet.sheet_view.showGridLines = False
    sheet.freeze_panes = "A4"

    start_date = date.fromisoformat(window_start)
    end_date = date.fromisoformat(window_end)
    display_end = end_date - timedelta(days=1)
    events = (
        pipeline_result.detailed if mode == "detailed" else pipeline_result.combined
    )
    events_by_date: dict[date, list[OperationalEvent]] = {}
    for event in events:
        events_by_date.setdefault(event.local_date, []).append(event)

    sheet.merge_cells("A1:C1")
    sheet["A1"] = "Weekly Function Sheet"
    sheet["A1"].font = Font(name="Arial", size=14, bold=True, color="17324D")
    sheet["A1"].alignment = Alignment(vertical="center")
    sheet.merge_cells("D1:K1")
    sheet["D1"] = f"{_date_label(start_date)} through {_date_label(display_end)}"
    sheet["D1"].font = Font(name="Arial", size=11, bold=True, color="334E68")
    sheet["D1"].alignment = Alignment(horizontal="center", vertical="center")
    sheet.row_dimensions[1].height = 28

    header_rule = Side(style="medium", color="17324D")
    for column_number, header in enumerate(FUNCTION_SHEET_HEADERS, start=1):
        cell = sheet.cell(row=3, column=column_number, value=header)
        cell.font = Font(
            name="Arial",
            size=11 if column_number <= 4 else 9,
            bold=True,
            color="17324D",
        )
        cell.alignment = Alignment(
            horizontal="left" if column_number <= 5 else "center",
            vertical="center",
        )
        cell.border = Border(bottom=header_rule)
    sheet.row_dimensions[3].height = 23

    current_row = 5
    current_date = start_date
    thin_rule = Side(style="thin", color="7B8794")
    medium_rule = Side(style="medium", color="52606D")
    day_fill = PatternFill("solid", fgColor="EAF0F6")
    review_fill = PatternFill("solid", fgColor="FFF4CE")

    while current_date < end_date:
        sheet.merge_cells(
            start_row=current_row,
            start_column=1,
            end_row=current_row,
            end_column=11,
        )
        day_cell = sheet.cell(row=current_row, column=1)
        day_cell.value = _date_label(current_date, include_weekday=True)
        day_cell.font = Font(name="Arial", size=10, bold=True, color="17324D")
        day_cell.fill = day_fill
        day_cell.alignment = Alignment(vertical="center")
        day_cell.border = Border(bottom=medium_rule)
        sheet.row_dimensions[current_row].height = 21
        current_row += 1

        day_events = sorted(
            events_by_date.get(current_date, []),
            key=lambda event: (
                event.start,
                event.space.casefold(),
                event.title.casefold(),
            ),
        )
        if not day_events:
            values = ["", "", "", "No scheduled functions", "", "", "", "", "", "", ""]
            day_events_for_style: list[OperationalEvent | None] = [None]
        else:
            values = []
            day_events_for_style = day_events

        for event in day_events_for_style:
            if event is not None:
                values = [
                    _time_label(event),
                    _safe_spreadsheet_text(event.space),
                    _safe_spreadsheet_text(event.group or event.title),
                    _safe_spreadsheet_text(event.function or event.title),
                    "",
                    "",
                    "",
                    "",
                    "",
                    "",
                    "",
                ]
            for column_number, value in enumerate(values, start=1):
                cell = sheet.cell(
                    row=current_row,
                    column=column_number,
                    value=value,
                )
                cell.font = Font(
                    name="Arial",
                    size=10 if column_number <= 5 else 9,
                    bold=column_number == 1 and event is not None,
                    italic=event is None,
                    color="52606D" if event is None else "000000",
                )
                cell.alignment = Alignment(
                    horizontal="center" if column_number >= 6 else "left",
                    vertical="top",
                    wrap_text=True,
                )
                cell.border = Border(
                    top=thin_rule,
                    bottom=medium_rule,
                    left=medium_rule if column_number == 1 else thin_rule,
                    right=medium_rule if column_number == 11 else thin_rule,
                )
                if event is not None and event.needs_review:
                    cell.fill = review_fill
            sheet.row_dimensions[current_row].height = 25
            current_row += 1
        current_row += 1
        current_date += timedelta(days=1)

    widths = {
        "A": 20,
        "B": 18,
        "C": 18,
        "D": 48,
        "E": 24,
        "F": 6,
        "G": 6,
        "H": 6,
        "I": 6,
        "J": 8,
        "K": 6,
    }
    for column, width in widths.items():
        sheet.column_dimensions[column].width = width

    sheet.page_setup.orientation = "landscape"
    sheet.page_setup.paperSize = sheet.PAPERSIZE_LETTER
    sheet.page_setup.fitToWidth = 1
    sheet.page_setup.fitToHeight = 0
    sheet.sheet_properties.pageSetUpPr.fitToPage = True
    sheet.page_margins = PageMargins(
        left=0.25,
        right=0.25,
        top=0.4,
        bottom=0.4,
        header=0.2,
        footer=0.2,
    )
    sheet.print_title_rows = "1:3"
    sheet.print_area = f"A1:K{max(current_row - 1, 5)}"


def _add_summary_sheet(
    workbook: Workbook,
    *,
    pipeline_result: PipelineResult,
    profile_name: str,
    mode: str,
    window_start: str,
    window_end: str,
) -> None:
    sheet = workbook.create_sheet("Summary")
    sheet.sheet_view.showGridLines = False
    sheet.merge_cells("A1:D1")
    sheet["A1"] = "VenueView Operational Review"
    sheet["A1"].font = Font(size=18, bold=True, color="FFFFFF")
    sheet["A1"].fill = PatternFill("solid", fgColor="163A5F")
    sheet["A1"].alignment = Alignment(vertical="center")
    sheet.row_dimensions[1].height = 32

    summary_rows = [
        ("Profile", _safe_spreadsheet_text(profile_name)),
        ("Output mode", mode.title()),
        ("First day included", window_start),
        ("Last day included", window_end),
        ("Detailed rows", len(pipeline_result.detailed)),
        ("Combined rows", len(pipeline_result.combined)),
        (
            "Rows needing review",
            sum(bool(event.needs_review) for event in pipeline_result.detailed),
        ),
        ("Excluded rows", pipeline_result.excluded_count),
        (
            "Source occurrences outside profile",
            pipeline_result.unassigned_source_count,
        ),
    ]
    for row_number, (label, value) in enumerate(summary_rows, start=3):
        sheet.cell(row=row_number, column=1, value=label)
        sheet.cell(row=row_number, column=2, value=value)
        sheet.cell(row=row_number, column=1).font = Font(bold=True, color="334E68")

    note_row = len(summary_rows) + 5
    sheet.merge_cells(
        start_row=note_row, start_column=1, end_row=note_row + 2, end_column=4
    )
    note = sheet.cell(row=note_row, column=1)
    note.value = (
        "Review before operational distribution. This locally generated workbook "
        "can contain calendar titles and group names. It intentionally omits "
        "calendar UIDs and is not yet an approved function-sheet template."
    )
    note.fill = PatternFill("solid", fgColor="FFF4CE")
    note.font = Font(color="7A4D00")
    note.alignment = Alignment(wrap_text=True, vertical="center")

    sheet.column_dimensions["A"].width = 36
    sheet.column_dimensions["B"].width = 28
    sheet.column_dimensions["C"].width = 16
    sheet.column_dimensions["D"].width = 16


def build_excel_download(
    *,
    pipeline_result: PipelineResult,
    profile_name: str,
    mode: str,
    window_start: str,
    window_end: str,
) -> DownloadArtifact:
    if mode not in {"detailed", "combined", "both"}:
        raise ValueError("Unsupported export mode")

    workbook = Workbook()
    workbook.properties.creator = "VenueView"
    workbook.properties.title = "VenueView Operational Review"
    workbook.properties.description = (
        "Locally generated operational review; source calendar identifiers omitted."
    )
    _add_function_sheet(
        workbook,
        pipeline_result=pipeline_result,
        mode=mode,
        window_start=window_start,
        window_end=window_end,
    )
    _add_summary_sheet(
        workbook,
        pipeline_result=pipeline_result,
        profile_name=profile_name,
        mode=mode,
        window_start=window_start,
        window_end=window_end,
    )

    if mode in {"detailed", "both"}:
        detailed_sheet = workbook.create_sheet("Detailed")
        _style_data_sheet(
            sheet=detailed_sheet,
            events=pipeline_result.detailed,
            table_name="DetailedEvents",
        )
    if mode in {"combined", "both"}:
        combined_sheet = workbook.create_sheet("Combined")
        _style_data_sheet(
            sheet=combined_sheet,
            events=pipeline_result.combined,
            table_name="CombinedEvents",
        )

    review_sheet = workbook.create_sheet("Review")
    _style_data_sheet(
        sheet=review_sheet,
        events=[event for event in pipeline_result.detailed if event.needs_review],
        table_name="ReviewEvents",
    )

    stream = BytesIO()
    workbook.save(stream)
    workbook.close()
    stem = _download_stem(
        profile_name=profile_name,
        window_start=window_start,
        window_end=window_end,
    )
    return DownloadArtifact(
        data=stream.getvalue(),
        filename=f"{stem}-review.xlsx",
        mimetype=("application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"),
    )
