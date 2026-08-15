from __future__ import annotations

from datetime import datetime, timedelta
from io import BytesIO
from zipfile import ZipFile
from zoneinfo import ZoneInfo

from openpyxl import load_workbook

from venueview.exporters import build_csv_download, build_excel_download
from venueview.ics import parse_ics
from venueview.models import OperationalEvent
from venueview.pipeline import PipelineResult, run_pipeline
from venueview.profiles import load_profile
from venueview.rules import load_rule_pack


def _pipeline_result(project_root):
    timezone = ZoneInfo("America/New_York")
    events = parse_ics(
        project_root / "data/synthetic/sample_calendar.ics",
        window_start=datetime(2026, 7, 17, tzinfo=timezone),
        window_end=datetime(2026, 7, 20, tzinfo=timezone),
    )
    profile = load_profile(project_root / "config/profiles/north_arena_rinks.json")
    rules = load_rule_pack(project_root / "config/rules/public_rules.json")
    return run_pipeline(events, profile, rules)


def _all_day_pipeline_result(project_root):
    timezone = ZoneInfo("America/New_York")
    events = parse_ics(
        project_root / "data/synthetic/sample_calendar.ics",
        window_start=datetime(2026, 7, 18, tzinfo=timezone),
        window_end=datetime(2026, 7, 19, tzinfo=timezone),
    )
    profile = load_profile(
        project_root / "config/profiles/aerial_training_park.json"
    )
    rules = load_rule_pack(project_root / "config/rules/public_rules.json")
    return run_pipeline(events, profile, rules)


def test_both_csv_export_is_an_in_memory_bundle_without_source_ids(project_root):
    artifact = build_csv_download(
        pipeline_result=_pipeline_result(project_root),
        profile_name="North Arena Rinks",
        mode="both",
        window_start="2026-07-17",
        window_end="2026-07-20",
    )

    assert artifact.filename == (
        "venueview-north-arena-rinks-2026-07-17-to-2026-07-20-csv.zip"
    )
    with ZipFile(BytesIO(artifact.data)) as archive:
        assert set(archive.namelist()) == {
            "README.txt",
            "combined.csv",
            "detailed.csv",
            "review.csv",
        }
        exported_text = "\n".join(
            archive.read(name).decode("utf-8-sig")
            for name in ("combined.csv", "detailed.csv", "review.csv")
        )
    assert "venueview-synthetic-" not in exported_text
    assert "Source Count" in exported_text


def test_excel_export_has_typed_review_sheets_and_no_source_ids(project_root):
    artifact = build_excel_download(
        pipeline_result=_pipeline_result(project_root),
        profile_name="North Arena Rinks",
        mode="both",
        window_start="2026-07-17",
        window_end="2026-07-20",
    )

    workbook = load_workbook(BytesIO(artifact.data), data_only=False)
    assert workbook.sheetnames == [
        "Function Sheet",
        "Summary",
        "Detailed",
        "Combined",
        "Review",
    ]
    assert workbook["Function Sheet"]["A1"].value == "Weekly Function Sheet"
    assert workbook["Function Sheet"]["A3"].value == "Time"
    assert workbook["Function Sheet"]["K3"].value == "Flip"
    assert workbook["Function Sheet"]["A5"].value == "Friday, July 17, 2026"
    assert workbook["Detailed"].freeze_panes == "A2"
    assert workbook["Detailed"]["A2"].number_format == "yyyy-mm-dd"
    assert workbook["Detailed"]["B2"].number_format == "hh:mm"
    assert workbook["Summary"]["B3"].value == "North Arena Rinks"
    assert workbook["Summary"]["A5"].value == "First day included"
    assert workbook["Summary"]["B5"].value == "2026-07-17"
    assert workbook["Summary"]["A6"].value == "Last day included"
    assert workbook["Summary"]["B6"].value == "2026-07-20"
    exported_values = "\n".join(
        str(cell.value)
        for sheet in workbook.worksheets
        for row in sheet.iter_rows()
        for cell in row
        if cell.value is not None
    )
    assert "venueview-synthetic-" not in exported_values
    assert "not yet an approved function-sheet template" in exported_values
    workbook.close()


def test_single_mode_export_only_adds_selected_data_sheet(project_root):
    artifact = build_excel_download(
        pipeline_result=_pipeline_result(project_root),
        profile_name="North Arena Rinks",
        mode="combined",
        window_start="2026-07-17",
        window_end="2026-07-20",
    )

    workbook = load_workbook(BytesIO(artifact.data), read_only=True)
    assert workbook.sheetnames == [
        "Function Sheet",
        "Summary",
        "Combined",
        "Review",
    ]
    workbook.close()


def test_all_day_exports_use_label_instead_of_midnight_times(project_root):
    result = _all_day_pipeline_result(project_root)
    csv_artifact = build_csv_download(
        pipeline_result=result,
        profile_name="Aerial Training Park",
        mode="detailed",
        window_start="2026-07-18",
        window_end="2026-07-19",
    )
    csv_text = csv_artifact.data.decode("utf-8-sig")
    assert "2026-07-18,All Day,," in csv_text
    assert "00:00" not in csv_text

    excel_artifact = build_excel_download(
        pipeline_result=result,
        profile_name="Aerial Training Park",
        mode="detailed",
        window_start="2026-07-18",
        window_end="2026-07-19",
    )
    workbook = load_workbook(BytesIO(excel_artifact.data), data_only=False)
    assert workbook["Detailed"]["B2"].value == "All Day"
    assert workbook["Detailed"]["C2"].value is None
    assert any(
        cell.value == "All Day"
        for row in workbook["Function Sheet"].iter_rows()
        for cell in row
    )
    workbook.close()


def test_exports_neutralize_formula_like_calendar_text():
    timezone = ZoneInfo("America/New_York")
    start = datetime(2026, 7, 17, 9, 0, tzinfo=timezone)
    event = OperationalEvent(
        source_uid="synthetic-formula-safety",
        title="=1+1",
        start=start,
        end=start + timedelta(hours=1),
        category_path="North Arena > Sports > Rinks > Rink B",
        venue="North Arena",
        space="Rink B",
        group="@Synthetic Group",
        function="Hockey",
    )
    result = PipelineResult(
        detailed=[event],
        combined=[event],
        excluded_count=0,
        unassigned_source_count=0,
    )

    csv_artifact = build_csv_download(
        pipeline_result=result,
        profile_name="North Arena Rinks",
        mode="combined",
        window_start="2026-07-17",
        window_end="2026-07-18",
    )
    csv_text = csv_artifact.data.decode("utf-8-sig")
    assert "'@Synthetic Group" in csv_text
    assert "'=1+1" in csv_text

    excel_artifact = build_excel_download(
        pipeline_result=result,
        profile_name="North Arena Rinks",
        mode="combined",
        window_start="2026-07-17",
        window_end="2026-07-18",
    )
    workbook = load_workbook(BytesIO(excel_artifact.data), data_only=False)
    assert workbook["Combined"]["F2"].value == "'@Synthetic Group"
    assert workbook["Combined"]["H2"].value == "'=1+1"
    assert workbook["Combined"]["H2"].data_type == "s"
    workbook.close()
