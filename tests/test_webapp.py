from io import BytesIO
import json
import os
import re
import stat
from zipfile import ZipFile

import pytest
from openpyxl import load_workbook

from venueview import webapp
from venueview.webapp import create_app


@pytest.fixture
def client(project_root, tmp_path):
    app = create_app(
        project_root / "config",
        private_rules_path=tmp_path / "private_rules.json",
    )
    app.config.update(TESTING=True)
    test_client = app.test_client()
    with test_client.session_transaction() as session:
        session["csrf_token"] = "test-csrf-token"
    return test_client


def form_data(project_root, **overrides):
    data = {
        "profile": "north_arena_rinks",
        "mode": "combined",
        "window_start": "2026-07-17",
        "window_end": "2026-07-20",
        "csrf_token": "test-csrf-token",
        "calendar": (
            BytesIO((project_root / "data/synthetic/sample_calendar.ics").read_bytes()),
            "synthetic.ics",
        ),
    }
    data.update(overrides)
    if data.pop("without_calendar", False):
        data.pop("calendar", None)
    return data


def test_health_endpoint_is_local_and_uncached(client):
    response = client.get("/health")

    assert response.status_code == 200
    assert response.json == {"status": "ok"}
    assert response.headers["Cache-Control"] == "no-store"
    assert response.headers["X-Content-Type-Options"] == "nosniff"
    assert response.headers["X-Frame-Options"] == "DENY"
    assert "frame-ancestors 'none'" in response.headers["Content-Security-Policy"]
    assert response.headers["Referrer-Policy"] == "no-referrer"


def test_home_page_lists_profiles_without_calendar_values(client):
    response = client.get("/")
    body = response.get_data(as_text=True)

    assert response.status_code == 200
    assert "Summit Events Center" in body
    assert "North Arena" in body
    assert "Combine Events" in body
    assert "Detailed Events" in body
    assert "Both Event Views" in body
    assert "Group one source event assigned to multiple locations" in body
    assert "Quit VenueView" in body
    assert "Standard public settings are active" in body
    assert "Import replacement settings" in body
    assert f"VenueView <span class=\"version\">v{webapp.__version__}" in body
    assert "Step 1: Organization settings" in body
    assert "Step 2: Choose calendar and dates" in body
    assert "How do I export an .ics calendar file?" in body
    assert "iCalendar</strong> or <strong>ICS" in body
    assert "Select the calendars or categories needed" in body
    assert "Do not share the feed link" in body
    assert '<script src="/runtime.js" defer></script>' in body
    assert "synthetic-fs-open" not in body
    cookie = client.get_cookie("session")
    assert cookie is not None
    assert cookie.http_only is True
    assert cookie.same_site == "Strict"


def test_private_rule_import_is_validated_persisted_and_activated(
    project_root, tmp_path
):
    private_path = tmp_path / "app-data" / "private_rules.json"
    app = create_app(project_root / "config", private_rules_path=private_path)
    app.config.update(TESTING=True)
    test_client = app.test_client()
    with test_client.session_transaction() as session:
        session["csrf_token"] = "test-csrf-token"
    payload = {
        "schema_version": 1,
        "metadata": {"name": "Synthetic private overlay"},
        "classification_rules": [
            {
                "id": "synthetic_private_group",
                "priority": 999,
                "field": "title",
                "operator": "contains",
                "value": "SYNTHETIC PRIVATE KEY",
                "assign": {"group": "Synthetic Private Group"},
            }
        ],
        "ignore_rules": [],
        "combination_rules": [],
    }

    response = test_client.post(
        "/operational-rules/import",
        data={
            "csrf_token": "test-csrf-token",
            "rules_file": (
                BytesIO(json.dumps(payload).encode("utf-8")),
                "private_rules.json",
            ),
        },
        content_type="multipart/form-data",
    )
    body = response.get_data(as_text=True)

    assert response.status_code == 200
    assert "checked, saved, and activated" in body
    assert "Using imported replacement settings" in body
    assert 'data-rules-source="imported"' in body
    assert "1 classification" in body
    assert json.loads(private_path.read_text(encoding="utf-8")) == payload
    if os.name != "nt":
        assert stat.S_IMODE(private_path.stat().st_mode) == 0o600


def test_private_edition_activates_bundled_rules_without_manual_import(
    project_root, tmp_path
):
    bundled_path = tmp_path / "bundle" / "approved_rules.json"
    bundled_path.parent.mkdir()
    bundled_path.write_text(
        json.dumps(
            {
                "schema_version": 1,
                "metadata": {"name": "Synthetic bundled private rules"},
                "classification_rules": [
                    {
                        "id": "synthetic_bundled_group",
                        "priority": 10,
                        "field": "title",
                        "operator": "contains",
                        "value": "SYNTHETIC BUNDLED KEY",
                        "assign": {"group": "Synthetic Bundled Group"},
                    }
                ],
                "ignore_rules": [],
                "combination_rules": [],
            }
        ),
        encoding="utf-8",
    )
    app = create_app(
        project_root / "config",
        private_rules_path=tmp_path / "user-data" / "private_rules.json",
        bundled_rules_path=bundled_path,
    )
    app.config.update(TESTING=True)

    body = app.test_client().get("/").get_data(as_text=True)

    assert "Using the approved built-in settings" in body
    assert 'data-rules-source="bundled"' in body
    assert "SYNTHETIC BUNDLED KEY" not in body


def test_imported_rules_replace_bundled_private_defaults(project_root, tmp_path):
    bundled_path = tmp_path / "bundle" / "approved_rules.json"
    bundled_path.parent.mkdir()
    bundled_path.write_text(
        json.dumps(
            {
                "schema_version": 1,
                "classification_rules": [],
                "ignore_rules": [],
                "combination_rules": [],
            }
        ),
        encoding="utf-8",
    )
    private_path = tmp_path / "user-data" / "private_rules.json"
    app = create_app(
        project_root / "config",
        private_rules_path=private_path,
        bundled_rules_path=bundled_path,
    )
    app.config.update(TESTING=True)
    test_client = app.test_client()
    with test_client.session_transaction() as session:
        session["csrf_token"] = "test-csrf-token"
    replacement = {
        "schema_version": 1,
        "metadata": {"name": "Synthetic imported replacement"},
        "classification_rules": [
            {
                "id": "synthetic_imported_group",
                "priority": 10,
                "field": "title",
                "operator": "contains",
                "value": "SYNTHETIC IMPORTED KEY",
                "assign": {"group": "Synthetic Imported Group"},
            }
        ],
        "ignore_rules": [],
        "combination_rules": [],
    }

    response = test_client.post(
        "/operational-rules/import",
        data={
            "csrf_token": "test-csrf-token",
            "rules_file": (
                BytesIO(json.dumps(replacement).encode("utf-8")),
                "replacement.json",
            ),
        },
        content_type="multipart/form-data",
    )
    body = response.get_data(as_text=True)

    assert response.status_code == 200
    assert "Using imported replacement settings" in body
    assert 'data-rules-source="imported"' in body
    assert json.loads(private_path.read_text(encoding="utf-8")) == replacement

    restarted = create_app(
        project_root / "config",
        private_rules_path=private_path,
        bundled_rules_path=bundled_path,
    )
    restarted.config.update(TESTING=True)
    restarted_body = restarted.test_client().get("/").get_data(as_text=True)
    assert 'data-rules-source="imported"' in restarted_body


def test_restore_defaults_removes_imported_settings_and_reactivates_bundle(
    project_root, tmp_path
):
    bundled_path = tmp_path / "bundle" / "approved_rules.json"
    bundled_path.parent.mkdir()
    bundled_path.write_text(
        json.dumps(
            {
                "schema_version": 1,
                "metadata": {"name": "Synthetic bundled settings"},
                "classification_rules": [],
                "ignore_rules": [],
                "combination_rules": [],
            }
        ),
        encoding="utf-8",
    )
    private_path = tmp_path / "user-data" / "private_rules.json"
    private_path.parent.mkdir()
    private_path.write_text(
        json.dumps(
            {
                "schema_version": 1,
                "metadata": {"name": "Synthetic replacement settings"},
                "classification_rules": [],
                "ignore_rules": [],
                "combination_rules": [],
            }
        ),
        encoding="utf-8",
    )
    app = create_app(
        project_root / "config",
        private_rules_path=private_path,
        bundled_rules_path=bundled_path,
    )
    app.config.update(TESTING=True)
    test_client = app.test_client()
    with test_client.session_transaction() as session:
        session["csrf_token"] = "test-csrf-token"

    response = test_client.post(
        "/operational-rules/restore-defaults",
        data={"csrf_token": "test-csrf-token"},
    )
    body = response.get_data(as_text=True)

    assert response.status_code == 200
    assert "Built-in organization settings were restored" in body
    assert 'data-rules-source="bundled"' in body
    assert "Using the approved built-in settings" in body
    assert not private_path.exists()


def test_invalid_private_rule_import_is_rejected_without_persisting_values(
    project_root, tmp_path
):
    private_path = tmp_path / "private_rules.json"
    app = create_app(project_root / "config", private_rules_path=private_path)
    app.config.update(TESTING=True)
    test_client = app.test_client()
    with test_client.session_transaction() as session:
        session["csrf_token"] = "test-csrf-token"
    private_value = "DO_NOT_RENDER_THIS_OPERATIONAL_VALUE"
    payload = {
        "schema_version": 1,
        "classification_rules": [
            {
                "id": "invalid_regex",
                "field": "title",
                "operator": "regex",
                "value": f"[{private_value}",
                "assign": {"group": "Private"},
            }
        ],
    }

    response = test_client.post(
        "/operational-rules/import",
        data={
            "csrf_token": "test-csrf-token",
            "rules_file": (
                BytesIO(json.dumps(payload).encode("utf-8")),
                "private_rules.json",
            ),
        },
        content_type="multipart/form-data",
    )
    body = response.get_data(as_text=True)

    assert response.status_code == 400
    assert "invalid regular expression" in body
    assert private_value not in body
    assert not private_path.exists()


def test_runtime_script_is_same_origin_and_contains_no_calendar_data(client):
    response = client.get("/runtime.js")
    body = response.get_data(as_text=True)

    assert response.status_code == 200
    assert response.mimetype == "application/javascript"
    assert 'fetch("/runtime/browser-opened"' in body
    assert 'navigator.sendBeacon("/runtime/browser-closed"' in body
    assert "Summit Skills - Open" not in body
    assert "script-src 'self'" in response.headers["Content-Security-Policy"]


def test_last_browser_page_closing_requests_clean_shutdown(client, monkeypatch):
    callbacks = []

    class ImmediateTimer:
        daemon = False

        def __init__(self, delay, callback, args=()):
            self.delay = delay
            self.callback = callback
            self.args = args
            self.cancelled = False

        def start(self):
            if not self.cancelled:
                self.callback(*self.args)

        def cancel(self):
            self.cancelled = True

    monkeypatch.setattr(webapp, "Timer", ImmediateTimer)
    client.application.config["VENUEVIEW_REQUEST_SHUTDOWN"] = lambda: callbacks.append(
        "shutdown"
    )

    first_open = client.post(
        "/runtime/browser-opened",
        data={"csrf_token": "test-csrf-token", "page_id": "page-one-123"},
    )
    second_open = client.post(
        "/runtime/browser-opened",
        data={"csrf_token": "test-csrf-token", "page_id": "page-two-456"},
    )
    first_close = client.post(
        "/runtime/browser-closed",
        data={"csrf_token": "test-csrf-token", "page_id": "page-one-123"},
    )

    assert first_open.status_code == 204
    assert second_open.status_code == 204
    assert first_close.status_code == 204
    assert callbacks == []

    second_close = client.post(
        "/runtime/browser-closed",
        data={"csrf_token": "test-csrf-token", "page_id": "page-two-456"},
    )

    assert second_close.status_code == 204
    assert callbacks == ["shutdown"]


def test_new_page_cancels_shutdown_during_browser_navigation(client, monkeypatch):
    callbacks = []
    timers = []

    class DeferredTimer:
        daemon = False

        def __init__(self, delay, callback, args=()):
            self.delay = delay
            self.callback = callback
            self.args = args
            self.cancelled = False
            timers.append(self)

        def start(self):
            return None

        def cancel(self):
            self.cancelled = True

        def fire(self):
            if not self.cancelled:
                self.callback(*self.args)

    monkeypatch.setattr(webapp, "Timer", DeferredTimer)
    client.application.config["VENUEVIEW_REQUEST_SHUTDOWN"] = lambda: callbacks.append(
        "shutdown"
    )

    client.post(
        "/runtime/browser-opened",
        data={"csrf_token": "test-csrf-token", "page_id": "old-page-123"},
    )
    client.post(
        "/runtime/browser-closed",
        data={"csrf_token": "test-csrf-token", "page_id": "old-page-123"},
    )
    assert len(timers) == 1
    assert timers[0].delay == webapp.BROWSER_CLOSE_GRACE_SECONDS

    client.post(
        "/runtime/browser-opened",
        data={"csrf_token": "test-csrf-token", "page_id": "new-page-456"},
    )
    timers[0].fire()

    assert timers[0].cancelled is True
    assert callbacks == []


def test_quit_button_requests_shutdown_and_requires_csrf(client, monkeypatch):
    callbacks = []

    class ImmediateTimer:
        daemon = False

        def __init__(self, _delay, callback, args=()):
            self.callback = callback
            self.args = args

        def start(self):
            self.callback(*self.args)

        def cancel(self):
            return None

    monkeypatch.setattr(webapp, "Timer", ImmediateTimer)
    client.application.config["VENUEVIEW_REQUEST_SHUTDOWN"] = lambda: callbacks.append(
        "shutdown"
    )

    rejected = client.post("/quit")
    accepted = client.post("/quit", data={"csrf_token": "test-csrf-token"})

    assert rejected.status_code == 400
    assert accepted.status_code == 200
    assert "VenueView has closed" in accepted.get_data(as_text=True)
    assert callbacks == ["shutdown"]


def test_safe_summary_does_not_render_event_titles(client, project_root):
    response = client.post(
        "/process",
        data=form_data(project_root),
        content_type="multipart/form-data",
    )
    body = response.get_data(as_text=True)

    assert response.status_code == 200
    assert "Processing summary" in body
    assert "calendar occurrences" in body
    assert "combined events" in body
    assert "Summit Skills - Open" not in body
    assert "venueview-synthetic-fs-open-1" not in body
    assert "Safe summary only" in body
    assert "Combination review" not in body


def test_dates_and_options_survive_missing_privacy_acknowledgement(
    client, project_root
):
    response = client.post(
        "/process",
        data=form_data(
            project_root,
            profile="summit_events_center",
            mode="both",
            window_start="2026-07-18",
            window_end="2026-07-22",
            group_multi_location="on",
            action="excel",
        ),
        content_type="multipart/form-data",
    )
    body = response.get_data(as_text=True)

    assert response.status_code == 400
    assert "Acknowledge the privacy notice" in body
    assert '<option value="summit_events_center" selected>' in body
    assert '<option value="both" selected>Both Event Views</option>' in body
    assert 'name="window_start" value="2026-07-18"' in body
    assert 'name="window_end" value="2026-07-22"' in body
    assert 'name="group_multi_location" checked' in body
    assert 'name="allow_sensitive_output" checked' not in body

    reopened = client.get("/").get_data(as_text=True)
    assert '<option value="summit_events_center" selected>' in reopened
    assert '<option value="both" selected>Both Event Views</option>' in reopened
    assert 'name="window_start" value="2026-07-18"' in reopened
    assert 'name="window_end" value="2026-07-22"' in reopened
    assert 'name="group_multi_location" checked' in reopened


def test_unchecked_options_replace_previously_remembered_choices(
    client, project_root
):
    first = client.post(
        "/process",
        data=form_data(
            project_root,
            group_multi_location="on",
            allow_sensitive_output="on",
        ),
        content_type="multipart/form-data",
    )
    assert first.status_code == 200

    changed = client.post(
        "/process",
        data=form_data(
            project_root,
            without_calendar=True,
            mode="detailed",
            window_start="2026-07-18",
            window_end="2026-07-19",
        ),
        content_type="multipart/form-data",
    )
    body = changed.get_data(as_text=True)

    assert changed.status_code == 200
    assert '<option value="detailed" selected>Detailed Events</option>' in body
    assert 'name="window_start" value="2026-07-18"' in body
    assert 'name="window_end" value="2026-07-19"' in body
    assert 'name="group_multi_location" checked' not in body
    assert 'name="allow_sensitive_output" checked' not in body


def test_form_actions_restore_the_submitted_scroll_position(client, project_root):
    response = client.post(
        "/process",
        data=form_data(project_root, scroll_position="742"),
        content_type="multipart/form-data",
    )
    runtime = client.get("/runtime.js").get_data(as_text=True)

    assert response.status_code == 200
    assert '<body data-restore-scroll="742">' in response.get_data(as_text=True)
    assert 'field.name = "scroll_position"' in runtime
    assert "requestAnimationFrame(() => scrollTo(0, position))" in runtime


def test_last_day_is_inclusive_and_support_summary_is_privacy_safe(
    client, project_root, tmp_path
):
    private_title = "SYNTHETIC LAST DAY PRIVATE TITLE"
    calendar = f"""BEGIN:VCALENDAR
VERSION:2.0
PRODID:-//VenueView//Inclusive Date Test//EN
BEGIN:VEVENT
UID:venueview-synthetic-inclusive-end
DTSTAMP:20260701T120000Z
DTSTART;TZID=America/New_York:20260720T100000
DTEND;TZID=America/New_York:20260720T110000
SUMMARY:{private_title}
CATEGORIES:Summit Events Center > Level 2 > Harbor Hall A
END:VEVENT
END:VCALENDAR
"""
    response = client.post(
        "/process",
        data={
            "profile": "summit_events_center",
            "mode": "detailed",
            "window_start": "2026-07-20",
            "window_end": "2026-07-20",
            "csrf_token": "test-csrf-token",
            "calendar": (BytesIO(calendar.encode("utf-8")), "private-name.ics"),
        },
        content_type="multipart/form-data",
    )
    body = response.get_data(as_text=True)

    assert response.status_code == 200
    assert "2026-07-20 through 2026-07-20 (inclusive)" in body
    assert "<strong>1</strong><span>calendar occurrences</span>" in body
    assert private_title not in body

    failed = client.post(
        "/process",
        data=form_data(
            project_root,
            without_calendar=True,
            window_start="not-a-date",
            window_end="2026-07-20",
        ),
        content_type="multipart/form-data",
    )
    assert failed.status_code == 400
    assert "Support code: VV-PROCESS-001" in failed.get_data(as_text=True)

    summary = client.get("/support-summary")
    summary_text = summary.get_data(as_text=True)

    assert summary.status_code == 200
    assert summary.mimetype == "text/plain"
    assert f"Version: {webapp.__version__}" in summary_text
    assert "Latest support code: VV-PROCESS-001" in summary_text
    assert "Calendar loaded in memory: Yes" in summary_text
    assert private_title not in summary_text
    assert "private-name.ics" not in summary_text
    assert str(tmp_path) not in summary_text


def test_acknowledged_local_preview_can_render_synthetic_titles(client, project_root):
    response = client.post(
        "/process",
        data=form_data(project_root, allow_sensitive_output="on"),
        content_type="multipart/form-data",
    )
    body = response.get_data(as_text=True)

    assert response.status_code == 200
    assert "Summit Skills Sessions" in body
    assert "Summit Skills - Open" in body
    assert "Combination review" in body
    assert "venueview-synthetic-fs-open-1" not in body


def test_combination_decision_is_reused_by_preview_and_excel(client, project_root):
    preview = client.post(
        "/process",
        data=form_data(project_root, allow_sensitive_output="on"),
        content_type="multipart/form-data",
    )
    body = preview.get_data(as_text=True)
    review_id_match = re.search(
        r'name="combination_id" value="([a-f0-9]{24})"', body
    )

    assert preview.status_code == 200
    assert review_id_match is not None
    assert "10 min gap" in body
    assert "Keep as one event" in body
    assert "Keep events separate" in body

    separated = client.post(
        "/process",
        data=form_data(
            project_root,
            without_calendar=True,
            action="separate_combination",
            combination_id=review_id_match.group(1),
            allow_sensitive_output="on",
        ),
        content_type="multipart/form-data",
    )
    separated_body = separated.get_data(as_text=True)

    assert separated.status_code == 200
    assert "<strong>6</strong><span>combined events</span>" in separated_body
    assert (
        'name="action" value="separate_combination" class="active" disabled'
        in separated_body
    )

    workbook_response = client.post(
        "/process",
        data=form_data(
            project_root,
            without_calendar=True,
            action="excel",
            allow_sensitive_output="on",
        ),
        content_type="multipart/form-data",
    )
    workbook = load_workbook(BytesIO(workbook_response.data), read_only=True)

    assert workbook_response.status_code == 200
    assert workbook["Combined"].max_row == 7
    workbook.close()

    recombined = client.post(
        "/process",
        data=form_data(
            project_root,
            without_calendar=True,
            action="keep_combined",
            combination_id=review_id_match.group(1),
            allow_sensitive_output="on",
        ),
        content_type="multipart/form-data",
    )

    assert recombined.status_code == 200
    assert "<strong>5</strong><span>combined events</span>" in recombined.get_data(
        as_text=True
    )


def test_new_upload_clears_combination_decisions(client, project_root):
    preview = client.post(
        "/process",
        data=form_data(project_root, allow_sensitive_output="on"),
        content_type="multipart/form-data",
    )
    review_id = re.search(
        r'name="combination_id" value="([a-f0-9]{24})"',
        preview.get_data(as_text=True),
    )
    assert review_id is not None

    separated = client.post(
        "/process",
        data=form_data(
            project_root,
            without_calendar=True,
            action="separate_combination",
            combination_id=review_id.group(1),
            allow_sensitive_output="on",
        ),
        content_type="multipart/form-data",
    )
    assert "<strong>6</strong><span>combined events</span>" in separated.get_data(
        as_text=True
    )

    replacement = client.post(
        "/process",
        data=form_data(project_root, allow_sensitive_output="on"),
        content_type="multipart/form-data",
    )
    assert "<strong>5</strong><span>combined events</span>" in replacement.get_data(
        as_text=True
    )


def test_invalid_request_does_not_echo_uploaded_filename(client, project_root):
    response = client.post(
        "/process",
        data=form_data(project_root, window_start="not-a-date"),
        content_type="multipart/form-data",
    )
    body = response.get_data(as_text=True)

    assert response.status_code == 400
    assert "could not process" in body
    assert "synthetic.ics" not in body


def test_post_requires_csrf_token(client, project_root):
    data = form_data(project_root)
    data.pop("csrf_token")
    response = client.post(
        "/process", data=data, content_type="multipart/form-data"
    )

    assert response.status_code == 400
    assert response.json["error"] == "The form security token is missing or invalid."


def test_calendar_can_be_reused_for_export_without_reupload(client, project_root):
    preview = client.post(
        "/process",
        data=form_data(project_root),
        content_type="multipart/form-data",
    )
    assert preview.status_code == 200
    assert "A calendar is loaded for this session" in preview.get_data(as_text=True)

    download = client.post(
        "/process",
        data=form_data(
            project_root,
            without_calendar=True,
            action="excel",
            allow_sensitive_output="on",
        ),
        content_type="multipart/form-data",
    )
    assert download.status_code == 200
    assert download.mimetype == (
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

    cleared = client.post(
        "/process",
        data=form_data(project_root, without_calendar=True, action="clear"),
        content_type="multipart/form-data",
    )
    assert cleared.status_code == 200
    assert "A calendar is loaded for this session" not in cleared.get_data(
        as_text=True
    )

    missing = client.post(
        "/process",
        data=form_data(project_root, without_calendar=True),
        content_type="multipart/form-data",
    )
    assert missing.status_code == 400
    assert "Choose an .ics calendar export" in missing.get_data(as_text=True)


def test_preview_can_group_striped_event_into_one_multi_location_event(
    client, project_root
):
    response = client.post(
        "/process",
        data=form_data(
            project_root,
            profile="summit_events_center",
            group_multi_location="on",
            allow_sensitive_output="on",
        ),
        content_type="multipart/form-data",
    )
    body = response.get_data(as_text=True)

    assert response.status_code == 200
    assert "Multi-location grouping: On" in body
    assert "Harbor Hall A / Harbor Hall B" in body
    assert 'name="group_multi_location" checked' in body


def test_all_day_preview_uses_label_instead_of_midnight_range(client, project_root):
    response = client.post(
        "/process",
        data=form_data(
            project_root,
            profile="aerial_training_park",
            mode="detailed",
            window_start="2026-07-18",
            window_end="2026-07-19",
            allow_sensitive_output="on",
        ),
        content_type="multipart/form-data",
    )
    body = response.get_data(as_text=True)

    assert response.status_code == 200
    assert "All Day" in body
    assert "00:00–00:00" not in body


def test_cross_origin_request_is_rejected(client):
    response = client.get("/", headers={"Origin": "https://evil.example"})

    assert response.status_code == 403
    assert response.json["error"] == "Cross-origin requests are not allowed."


def test_extension_origin_is_accepted_with_valid_session_csrf(client, project_root):
    response = client.post(
        "/process",
        data=form_data(project_root),
        content_type="multipart/form-data",
        headers={"Origin": "chrome-extension://synthetic-extension"},
    )

    assert response.status_code == 200
    assert "Processing summary" in response.get_data(as_text=True)


def test_extension_origin_without_valid_csrf_remains_blocked(client, project_root):
    data = form_data(project_root)
    data["csrf_token"] = "wrong-token"
    response = client.post(
        "/process",
        data=data,
        content_type="multipart/form-data",
        headers={"Origin": "chrome-extension://synthetic-extension"},
    )

    assert response.status_code == 403
    assert response.json["error"] == "Cross-origin requests are not allowed."


def test_download_requires_explicit_privacy_acknowledgement(client, project_root):
    response = client.post(
        "/process",
        data=form_data(project_root, action="excel"),
        content_type="multipart/form-data",
    )
    body = response.get_data(as_text=True)

    assert response.status_code == 400
    assert "Acknowledge the privacy notice" in body
    assert "synthetic.ics" not in body


def test_combined_csv_download_is_uncached_and_omits_source_uid(client, project_root):
    response = client.post(
        "/process",
        data=form_data(
            project_root,
            action="csv",
            allow_sensitive_output="on",
        ),
        content_type="multipart/form-data",
    )
    body = response.get_data(as_text=True)

    assert response.status_code == 200
    assert response.mimetype == "text/csv"
    assert response.headers["Cache-Control"] == "no-store"
    assert response.headers["Content-Disposition"].endswith(
        "venueview-north-arena-rinks-2026-07-17-to-2026-07-20-combined.csv"
    )
    assert "Date,Start,End,Venue,Space,Group,Function,Title" in body
    assert "venueview-synthetic-" not in body


def test_both_csv_download_returns_three_operational_views(client, project_root):
    response = client.post(
        "/process",
        data=form_data(
            project_root,
            mode="both",
            action="csv",
            allow_sensitive_output="on",
        ),
        content_type="multipart/form-data",
    )

    assert response.status_code == 200
    assert response.mimetype == "application/zip"
    with ZipFile(BytesIO(response.data)) as archive:
        assert {"detailed.csv", "combined.csv", "review.csv"}.issubset(
            archive.namelist()
        )


def test_excel_download_opens_as_a_review_workbook(client, project_root):
    response = client.post(
        "/process",
        data=form_data(
            project_root,
            mode="both",
            action="excel",
            allow_sensitive_output="on",
        ),
        content_type="multipart/form-data",
    )

    assert response.status_code == 200
    assert response.mimetype == (
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    workbook = load_workbook(BytesIO(response.data), read_only=True)
    assert workbook.sheetnames == [
        "Function Sheet",
        "Summary",
        "Detailed",
        "Combined",
        "Review",
    ]
    workbook.close()


def test_browser_open_is_delayed_until_the_local_service_can_start(monkeypatch):
    captured = {}

    class FakeTimer:
        daemon = False

        def __init__(self, delay, callback, args):
            captured.update(delay=delay, callback=callback, args=args)

        def start(self):
            captured["started"] = True

    monkeypatch.setattr(webapp, "Timer", FakeTimer)
    timer = webapp._schedule_browser_open("http://127.0.0.1:8765")

    assert captured["delay"] == 0.8
    assert captured["callback"] is webapp.webbrowser.open_new_tab
    assert captured["args"] == ("http://127.0.0.1:8765",)
    assert captured["started"] is True
    assert timer.daemon is True
